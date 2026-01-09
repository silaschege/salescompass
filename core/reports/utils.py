import csv
import json
from django.http import HttpResponse
from django.db.models import Q, Sum, Avg, Count, F
from django.db.models.functions import TruncMonth, TruncYear, TruncDay
from django.utils import timezone
from datetime import datetime, timedelta
from django.apps import apps
 
def generate_report(report_id, export_format='csv', user=None):
    """
    Generate report in specified format.
    """
    from .models import Report
    report = Report.objects.get(id=report_id)
    query_config = report.query_config or {}
    
    # Build queryset based on report type and config
    if report.report_type == 'sales_performance':
        queryset = _build_sales_performance_query(query_config)
    elif report.report_type == 'esg_impact':
        queryset = _build_esg_impact_query(query_config)
    elif report.report_type == 'pipeline_forecast':
        queryset = _build_pipeline_forecast_query(query_config)
    elif report.report_type == 'csrd_compliance':
        queryset = _build_csrd_compliance_query(query_config)
    elif report.report_type == 'leads_recent':
        queryset = _build_leads_recent_query(query_config)
    elif report.report_type == 'cases_recent':
        queryset = _build_cases_recent_query(query_config)
    else:  # custom
        queryset = _build_custom_query(query_config)

    # Apply Joins (Cross-object)
    joins = query_config.get('joins', [])
    for join in joins:
        queryset = queryset.select_related(join) if '__' not in join else queryset.prefetch_related(join)

    # Apply Formulas (Computed fields)
    formulas = query_config.get('formulas', [])
    for formula in formulas:
        name = formula.get('name')
        expression = formula.get('expression')
        if name and expression:
            try:
                # Enhanced parser for expressions using F objects
                from django.db.models import FloatField, ExpressionWrapper
                import re

                def to_expression(expr_str):
                    tokens = re.split(r'(\s*[\+\-\*/\(\)]\s*)', expr_str)
                    processed_tokens = []
                    for t in tokens:
                        t = t.strip()
                        if not t: continue
                        if t in '+-*/()':
                            processed_tokens.append(t)
                        else:
                            try:
                                float(t)
                                processed_tokens.append(t)
                            except ValueError:
                                processed_tokens.append(f"F('{t}')")
                    return "".join(processed_tokens)

                python_expr = to_expression(expression)
                queryset = queryset.annotate(**{name: ExpressionWrapper(eval(python_expr), output_field=FloatField())})
            except Exception as e:
                print(f"Formula error: {e}")
                pass

    # Apply filters from query_config
    if 'filters' in query_config:
        queryset = _apply_filters(queryset, query_config['filters'])
    
    # Apply ordering
    if 'sort_by' in query_config:
        queryset = queryset.order_by(query_config['sort_by'])
    else:
        queryset = queryset.order_by('-report_created_at')
    
    # Get fields to include
    fields = query_config.get('fields', _get_default_fields(report.report_type))

    # Apply Conditional Formatting (Annotates data for UI)
    formatting = query_config.get('formatting', [])
    data_rows = list(queryset) # Evaluate for formatting
    _apply_formatting(data_rows, formatting)
    
    # Generate export
    if export_format == 'csv':
        return _generate_csv_export(data_rows, fields, report.report_type)
    elif export_format == 'xlsx':
        return _generate_xlsx_export(data_rows, fields, report.report_type)
    else:  # pdf
        return _generate_pdf_export(data_rows, fields, report.report_type)


def _get_default_fields(report_type):
    """Get default fields based on report type."""
    defaults = {
        'sales_performance': ['account__name', 'name', 'amount', 'close_date', 'stage'],
        'esg_impact': ['name', 'industry', 'esg_engagement', 'health_score'],
        'pipeline_forecast': ['account__name', 'name', 'amount', 'stage', 'weighted_value'],
        'csrd_compliance': ['name', 'industry', 'esg_score__csr_ready', 'esg_score__overall_esg_score'],
        'leads_recent': ['first_name', 'last_name', 'company', 'lead_score', 'created_at', 'status'],
        'cases_recent': ['subject', 'account__name', 'priority', 'status', 'created_at', 'resolved_at'],
        'custom': ['id', 'name']
    }
    return defaults.get(report_type, ['id', 'name'])


def _build_sales_performance_query(config):
    """Build sales performance report query."""
    try:
        from opportunities.models import Opportunity
        return Opportunity.objects.select_related('account', 'owner').filter(
            stage='closed_won'
        )
    except ImportError:
        # Fallback if opportunities app is not available
        return Opportunity.objects.none()


def _build_esg_impact_query(config):
    """Build ESG impact report query."""
    try:
        from accounts.models import Account
        return Account.objects.select_related('esg_score').filter(
            esg_score__isnull=False
        )
    except ImportError:
        return Account.objects.none()


def _build_pipeline_forecast_query(config):
    """Build pipeline forecast report query."""
    try:
        from opportunities.models import Opportunity
        return Opportunity.objects.select_related('account').filter(
            stage__in=['prospecting', 'qualification', 'proposal', 'negotiation']
        )
    except ImportError:
        return Opportunity.objects.none()


def _build_csrd_compliance_query(config):
    """Build CSRD compliance report query."""
    try:
        from accounts.models import Account
        return Account.objects.filter(
            esg_score__csrd_ready=True
        )
    except ImportError:
        return Account.objects.none()


def _build_leads_recent_query(config):
    """Build leads from last 7 days query."""
    try:
        from leads.models import Lead
        from django.utils import timezone
        from datetime import timedelta
        seven_days_ago = timezone.now() - timedelta(days=7)
        return Lead.objects.filter(created_at__gte=seven_days_ago)
    except ImportError:
        return Lead.objects.none()


def _build_cases_recent_query(config):
    """Build cases from last 3 months query."""
    try:
        from cases.models import Case
        from django.utils import timezone
        from datetime import timedelta
        three_months_ago = timezone.now() - timedelta(days=90)
        return Case.objects.select_related('account').filter(created_at__gte=three_months_ago)
    except ImportError:
        return Case.objects.none()


def _build_custom_query(config):
    """Build custom report query based on entities specified."""
    entity = config.get('entity', 'account') # Single primary entity
    
    model_map = {
        'account': 'accounts.Account',
        'opportunity': 'opportunities.Opportunity',
        'lead': 'leads.Lead',
        'case': 'support.Case',
        'task': 'tasks.Task',
    }

    model_path = model_map.get(entity, 'accounts.Account')
    app_label, model_name = model_path.split('.')
    Model = apps.get_model(app_label, model_name)
    
    queryset = Model.objects.all()

    # Apply performance optimizations (Select Related / Prefetch Related)
    joins = config.get('joins', [])
    select_related_list = []
    prefetch_related_list = []
    
    for join in joins:
        if isinstance(join, str):
            if '__' in join: # Likely a nested prefetch
                prefetch_related_list.append(join)
            else:
                select_related_list.append(join)
    
    if select_related_list:
        queryset = queryset.select_related(*select_related_list)
    if prefetch_related_list:
        queryset = queryset.prefetch_related(*prefetch_related_list)

    return queryset


def _apply_formatting(data_rows, formatting_rules):
    """
    Apply conditional formatting rules to data rows.
    Adds a '_formatting' dict to each object in the list.
    """
    if not formatting_rules:
        return

    for obj in data_rows:
        obj._formatting = {}
        for rule in formatting_rules:
            field = rule.get('field')
            operator = rule.get('operator')
            value = rule.get('value')
            style = rule.get('style') # e.g. {'color': 'red', 'font-weight': 'bold'}
            
            if not (field and operator and style):
                continue
            
            field_val = getattr(obj, field, None)
            if field_val is None:
                continue

            match = False
            if operator == 'gt': match = field_val > value
            elif operator == 'lt': match = field_val < value
            elif operator == 'eq': match = field_val == value
            elif operator == 'contains': match = str(value) in str(field_val)
            
            if match:
                obj._formatting[field] = style
    """Apply filters to queryset (handles both dict and list of dicts)."""
    if isinstance(filters, list):
        for f in filters:
            field = f.get('field')
            operator = f.get('operator')
            value = f.get('value')
            if field and operator and value:
                lookup = f"{field}__{operator}" if operator != 'exact' else field
                queryset = queryset.filter(**{lookup: value})
    else:
        for field, value in filters.items():
            if isinstance(value, list):
                queryset = queryset.filter(**{f"{field}__in": value})
            elif isinstance(value, dict):
                for op, val in value.items():
                    lookup = f"{field}{op}" if op.startswith('__') else f"{field}__{op}"
                    queryset = queryset.filter(**{lookup: val})
            else:
                queryset = queryset.filter(**{field: value})
    return queryset


def get_report_data(query_config):
    """
    Generate data for charts/tables based on query_config.
    Returns: { 'labels': [...], 'datasets': [{ 'label': '...', 'data': [...] }] }
    """
    entity = query_config.get('entity', 'account')
    filters = query_config.get('filters', [])
    group_by = query_config.get('group_by')
    chart_type = query_config.get('chart_type', 'bar')
    formulas = query_config.get('formulas', [])
    pivot_config = query_config.get('pivot_config', {})

    # Map entity to model
    model_map = {
        'account': 'accounts.Account',
        'opportunity': 'opportunities.Opportunity',
        'lead': 'leads.Lead',
        'case': 'cases.Case', # Fixed mapping from support.Case to cases.Case
        'task': 'tasks.Task',
    }

    model_path = model_map.get(entity)
    if not model_path:
        return {'error': 'Invalid entity'}

    try:
        app_label, model_name = model_path.split('.')
        Model = apps.get_model(app_label, model_name)
    except LookupError:
        return {'error': f'Model {entity} not found'}

    # Base QuerySet
    queryset = Model.objects.all()

    # Apply Formulas
    for formula in formulas:
        name = formula.get('name')
        expr = formula.get('expression')
        if name and expr:
            try:
                from django.db.models import FloatField, ExpressionWrapper
                import re
                def to_expression(expr_str):
                    tokens = re.split(r'(\s*[\+\-\*/\(\)]\s*)', expr_str)
                    processed_tokens = []
                    for t in tokens:
                        t = t.strip()
                        if not t: continue
                        if t in '+-*/()':
                            processed_tokens.append(t)
                        else:
                            try:
                                float(t)
                                processed_tokens.append(t)
                            except ValueError:
                                processed_tokens.append(f"F('{t}')")
                    return "".join(processed_tokens)
                
                python_expr = to_expression(expr)
                queryset = queryset.annotate(**{name: ExpressionWrapper(eval(python_expr), output_field=FloatField())})
            except Exception:
                pass

    # Aggregation
    labels = []
    datasets = []

    if chart_type == 'pivot' and pivot_config.get('rows') and pivot_config.get('cols'):
        row_field = pivot_config['rows']
        col_field = pivot_config['cols']
        
        # Aggregate by both fields
        data = queryset.values(row_field, col_field).annotate(count=Count('id')).order_by(row_field, col_field)
        
        # Group into a matrix
        rows = sorted(list(set(item[row_field] for item in data if item[row_field])))
        cols = sorted(list(set(item[col_field] for item in data if item[col_field])))
        
        labels = [str(c) for c in cols]
        for r in rows:
            row_data = []
            for c in cols:
                count = next((item['count'] for item in data if item[row_field] == r and item[col_field] == c), 0)
                row_data.append(count)
            
            datasets.append({
                'label': str(r),
                'data': row_data
            })
            
        return {
            'labels': labels,
            'datasets': datasets,
            'is_pivot': True
        }

    if group_by:
        # Check for date truncation
        date_field = None
        trunc_func = None

        if '__month' in group_by:
            date_field = group_by.replace('__month', '')
            trunc_func = TruncMonth
        elif '__year' in group_by:
            date_field = group_by.replace('__year', '')
            trunc_func = TruncYear
        elif '__day' in group_by:
            date_field = group_by.replace('__day', '')
            trunc_func = TruncDay

        if date_field and trunc_func:
            data = queryset.annotate(period=trunc_func(date_field)).values('period').annotate(count=Count('id')).order_by('period')
            for item in data:
                labels.append(item['period'].strftime('%Y-%m-%d') if item['period'] else 'None')
                datasets.append(item['count'])
        else:
            # Special case for geographic mapping
            is_geo = group_by in ['billing_country', 'territory', 'shipping_country', 'billing_state', 'shipping_state']
            if is_geo:
                chart_type = 'choropleth'
            
            # Categorical grouping with optional secondary measure
            measure = query_config.get('measure', 'count')
            if measure == 'sum' and query_config.get('measure_field'):
                data = queryset.values(group_by).annotate(value=Sum(query_config['measure_field'])).order_by('-value')
            elif measure == 'avg' and query_config.get('measure_field'):
                data = queryset.values(group_by).annotate(value=Avg(query_config['measure_field'])).order_by('-value')
            else:
                data = queryset.values(group_by).annotate(value=Count('id')).order_by('-value')

            for item in data:
                labels.append(str(item[group_by]) if item[group_by] else 'None')
                datasets.append(item['value'])
    else:
        # No grouping, maybe just top 10?
        labels = ["Total Records"]
        datasets = [queryset.count()]

    # Standardize result for Chart.js
    if not isinstance(datasets, list) or (datasets and not isinstance(datasets[0], dict)):
        datasets = [{
            'label': entity.title(),
            'data': datasets,
            'backgroundColor': 'rgba(111, 66, 193, 0.5)',
            'borderColor': 'rgba(111, 66, 193, 1)',
            'borderWidth': 1
        }]

    return {
        'labels': labels,
        'datasets': datasets,
        'chart_type': chart_type
    }


def _generate_csv_export(queryset, fields, report_type):
    """Generate CSV export."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = []
    for field in fields:
        headers.append(field.replace('__', ' ').replace('_', ' ').title())
    writer.writerow(headers)
    
    # Write data
    for obj in queryset:
        row = []
        for field in fields:
            if hasattr(obj, field):
                value = getattr(obj, field)
                if value is None:
                    row.append('')
                elif callable(value):
                    row.append(str(value()))
                else:
                    row.append(str(value))
            else:
                # Handle related fields (e.g., 'account__name')
                if '__' in field:
                    try:
                        value = obj
                        for f in field.split('__'):
                            value = getattr(value, f)
                        if value is None:
                            row.append('')
                        elif callable(value):
                            row.append(str(value()))
                        else:
                            row.append(str(value))
                    except (AttributeError, TypeError):
                        row.append('')
                else:
                    row.append('')
        writer.writerow(row)
    
    return response


def _generate_xlsx_export(queryset, fields, report_type):
    """Generate Excel export."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Write headers
        headers = []
        for field in fields:
            headers.append(field.replace('__', ' ').replace('_', ' ').title())
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(fields, 1):
                if hasattr(obj, field):
                    value = getattr(obj, field)
                    if value is None:
                        ws.cell(row=row_num, column=col_num, value='')
                    elif callable(value):
                        ws.cell(row=row_num, column=col_num, value=str(value()))
                    else:
                        ws.cell(row=row_num, column=col_num, value=value)
                else:
                    # Handle related fields
                    if '__' in field:
                        try:
                            value = obj
                            for f in field.split('__'):
                                value = getattr(value, f)
                            if value is None:
                                ws.cell(row=row_num, column=col_num, value='')
                            elif callable(value):
                                ws.cell(row=row_num, column=col_num, value=str(value()))
                            else:
                                ws.cell(row=row_num, column=col_num, value=value)
                        except (AttributeError, TypeError):
                            ws.cell(row=row_num, column=col_num, value='')
                    else:
                        ws.cell(row=row_num, column=col_num, value='')
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        # Fallback to CSV if openpyxl is not installed
        return _generate_csv_export(queryset, fields, report_type)


def _generate_pdf_export(queryset, fields, report_type):
    """Generate PDF export."""
    try:
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(f"SalesCompass Report - {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Build table data
        table_data = []
        
        # Headers
        headers = []
        for field in fields:
            headers.append(field.replace('__', ' ').replace('_', ' ').title())
        table_data.append(headers)
        
        # Data rows
        for obj in queryset[:1000]:  # Limit to 1000 rows for PDF
            row = []
            for field in fields:
                if hasattr(obj, field):
                    value = getattr(obj, field)
                    if value is None:
                        row.append('')
                    elif callable(value):
                        row.append(str(value()))
                    else:
                        row.append(str(value))
                else:
                    # Handle related fields
                    if '__' in field:
                        try:
                            value = obj
                            for f in field.split('__'):
                                value = getattr(value, f)
                            if value is None:
                                row.append('')
                            elif callable(value):
                                row.append(str(value()))
                            else:
                                row.append(str(value))
                        except (AttributeError, TypeError):
                            row.append('')
                    else:
                        row.append('')
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        # Create response
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
        
    except ImportError:
        # Fallback to CSV if reportlab is not installed
        return _generate_csv_export(queryset, fields, report_type)


def create_scheduled_report_export(schedule_id):
    """
    Create a scheduled report export (for Celery task).
    """
    from .models import ReportSchedule, ReportExport
    schedule = ReportSchedule.objects.get(id=schedule_id)
    
    try:
        # Generate the report
        response = generate_report(schedule.report.id, schedule.export_format)
        
        # Save to file field
        export = ReportExport.objects.create(
            report=schedule.report,
            export_format=schedule.export_format,
            status='completed',
            tenant_id=schedule.tenant_id
        )
        
        # Save the file
        filename = f"report_{schedule.report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{schedule.export_format}"
        export.file.save(filename, response)
        export.save()
        
        # Send email to recipients
        from .utils import send_scheduled_report_email
        send_scheduled_report_email(export, schedule.recipients)
        
        # Schedule next run
        _schedule_next_run(schedule)
        
    except Exception as e:
        # Log error and create failed export record
        ReportExport.objects.create(
            report=schedule.report,
            export_format=schedule.export_format,
            status='failed',
            error_message=str(e),
            tenant_id=schedule.tenant_id
        )


def send_scheduled_report_email(export, recipients):
    """
    Send scheduled report email to recipients.
    """
    from django.core.mail import EmailMultiAlternatives
    from django.template.loader import render_to_string
    
    html_content = render_to_string('reports/export_email.html', {
        'export': export,
        'report': export.report
    })
    
    msg = EmailMultiAlternatives(
        subject=f"SalesCompass Report: {export.report.report_name}",
        body="Please find your scheduled report attached.",
        from_email="reports@salescompass.com",
        to=recipients
    )
    msg.attach_alternative(html_content, "text/html")
    
    # Attach the file
    if export.file:
        msg.attach_file(export.file.path)
    
    msg.send()


def _schedule_next_run(schedule):
    """
    Schedule the next run for a report schedule.
    """
    from datetime import timedelta
    now = timezone.now()
    
    if schedule.frequency == 'daily':
        schedule.next_run = now + timedelta(days=1)
    elif schedule.frequency == 'weekly':
        schedule.next_run = now + timedelta(weeks=1)
    elif schedule.frequency == 'monthly':
        schedule.next_run = now + timedelta(days=30)
    else:  # quarterly
        schedule.next_run = now + timedelta(days=90)
    
    schedule.save(update_fields=['next_run'])


def get_dashboard_widget_data(widget_id):
    """
    Get data for a dashboard widget.
    """
    from .models import DashboardWidget
    widget = DashboardWidget.objects.get(id=widget_id)
    report = widget.report
    
    # Generate the data based on report query_config
    query_config = report.query_config or {}
    fields = query_config.get('fields', _get_default_fields(report.report_type))
    
    # Build queryset
    if report.report_type == 'sales_performance':
        queryset = _build_sales_performance_query(query_config)
    elif report.report_type == 'esg_impact':
        queryset = _build_esg_impact_query(query_config)
    elif report.report_type == 'pipeline_forecast':
        queryset = _build_pipeline_forecast_query(query_config)
    elif report.report_type == 'csrd_compliance':
        queryset = _build_csrd_compliance_query(query_config)
    else:
        queryset = _build_custom_query(query_config)
    
    # Apply filters
    if 'filters' in query_config:
        queryset = _apply_filters(queryset, query_config['filters'])
    
    # Get data based on widget type
    if widget.widget_type == 'kpi_card':
        return {
            'value': queryset.count(),
            'label': report.report_name
        }
    elif widget.widget_type == 'chart':
        # Return chart data
        return {
            'labels': [],
            'data': [],
            'type': 'bar'
        }
    elif widget.widget_type == 'table':
        # Return table data
        data = []
        headers = []
        for field in fields:
            headers.append(field.replace('__', ' ').replace('_', ' ').title())
        
        for obj in queryset[:10]:  # Limit to 10 rows for dashboard
            row = []
            for field in fields:
                if hasattr(obj, field):
                    value = getattr(obj, field)
                    row.append(str(value) if value is not None else '')
                else:
                    row.append('')
            data.append(row)
        
        return {
            'headers': headers,
            'rows': data
        }
    else:  # trend
        return {
            'labels': [],
            'data': []
        }